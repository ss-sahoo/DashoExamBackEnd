"""
Timetable management APIs - Super Admin creates centers, admins, teachers, students.
All these endpoints require Super Admin authentication.
"""

from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db import transaction
from .models import Institute, Center, Program, Batch, Enrollment, User
from .utils import generate_user_code, generate_password


def _check_super_admin(request):
    """Helper to check if user is super admin."""
    if not request.user.is_authenticated:
        return False, Response(
            {"detail": "Authentication required."},
            status=status.HTTP_401_UNAUTHORIZED,
        )
    if request.user.role not in ['super_admin', 'SUPER_ADMIN']:
        return False, Response(
            {"detail": "Only Super Admin can perform this action."},
            status=status.HTTP_403_FORBIDDEN,
        )
    return True, None


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_center(request):
    """
    Super Admin creates a new Center.
    
    Payload:
    {
        "institute_id": 1,  # Institute ID (required)
        "name": "Allen - Jaipur Center",
        "city": "Jaipur",
        "address": "Optional address"
    }
    """
    is_super, error_response = _check_super_admin(request)
    if not is_super:
        return error_response
    
    institute_id = request.data.get("institute_id")
    name = request.data.get("name")
    city = request.data.get("city")
    address = request.data.get("address", "")
    
    if not institute_id or not name or not city:
        return Response(
            {"detail": "institute_id, name, and city are required."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    # Find institute by ID
    try:
        institute = Institute.objects.get(id=institute_id)
    except Institute.DoesNotExist:
        return Response(
            {"detail": f"Institute with id '{institute_id}' not found."},
            status=status.HTTP_404_NOT_FOUND,
        )
    
    try:
        center = Center.objects.create(
            institute=institute,
            name=name,
            city=city,
            address=address,
        )
        return Response(
            {
                "id": str(center.id),
                "name": center.name,
                "city": center.city,
                "institute": {
                    "id": institute.id,
                    "name": institute.name,
                },
            },
            status=status.HTTP_201_CREATED,
        )
    except Exception as e:
        return Response(
            {"detail": f"Error creating center: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_center(request, center_id):
    """
    Super Admin updates a Center.
    
    Payload:
    {
        "name": "Updated Name",
        "city": "Updated City",
        "address": "Updated Address"
    }
    """
    is_super, error_response = _check_super_admin(request)
    if not is_super:
        return error_response
    
    try:
        center = Center.objects.get(id=center_id)
    except Center.DoesNotExist:
        return Response(
            {"detail": f"Center with id '{center_id}' not found."},
            status=status.HTTP_404_NOT_FOUND,
        )
    
    name = request.data.get("name")
    city = request.data.get("city")
    address = request.data.get("address")
    
    if name:
        center.name = name
    if city:
        center.city = city
    if address is not None:
        center.address = address
        
    try:
        center.save()
        return Response(
            {
                "id": str(center.id),
                "name": center.name,
                "city": center.city,
                "address": center.address,
                "institute": {
                    "id": center.institute.id,
                    "name": center.institute.name,
                },
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"detail": f"Error updating center: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_center(request, center_id):
    """
    Super Admin deletes a Center.
    """
    is_super, error_response = _check_super_admin(request)
    if not is_super:
        return error_response
    
    try:
        center = Center.objects.get(id=center_id)
    except Center.DoesNotExist:
        return Response(
            {"detail": f"Center with id '{center_id}' not found."},
            status=status.HTTP_404_NOT_FOUND,
        )
    
    try:
        center.delete()
        return Response(
            {"message": "Center deleted successfully."},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"detail": f"Error deleting center: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_admin(request):
    """
    Super Admin creates a new Admin user for a center.
    
    Auto-generates:
    - username (code): ADM-<center_code>-<3_digits>
    - password: Admin@<center_code><year>
    
    Payload:
    {
        "center_name": "Allen - Jaipur Center",  # OR use "center_id": "uuid"
        "name": "Admin Name",
        "email": "admin@example.com",
        "phone_number": "9876543210"
    }
    
    Returns:
    {
        "username": "ADM-XXXX-123",
        "password": "Admin@XXXX2025",
        "user_id": "uuid"
    }
    """
    is_super, error_response = _check_super_admin(request)
    if not is_super:
        return error_response
    
    center_name = request.data.get("center_name")
    center_id = request.data.get("center_id")
    name = request.data.get("name")
    email = request.data.get("email", "")
    phone_number = request.data.get("phone_number", "")
    
    if not name:
        return Response(
            {"detail": "name is required."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    # Find center by ID (preferred) or name
    center, error_response = _get_center_by_name_or_id(center_name=center_name, center_id=center_id)
    if error_response:
        return error_response
    
    # Generate code and password - use 'ADMIN' role for timetable system
    center_code = center.name[:4].replace(" ", "").replace("-", "")
    username = generate_user_code('ADMIN', center_code)
    password = generate_password('ADMIN', center_code)
    
    # Split name
    name_parts = name.strip().split()
    first_name = name_parts[0] if name_parts else name
    last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
    
    try:
        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                email=email or f"{username}@temp.com",
                password=password,
                first_name=first_name,
                last_name=last_name,
                phone=phone_number or "",
                phone_number=phone_number or "",
                role='admin',  # Use lowercase admin role
                center=center,
                institute=center.institute,
            )
            
            # Add admin to center's admins
            center.admins.add(user)
            
            return Response(
                {
                    "message": "Admin created successfully.",
                    "username": username,
                    "password": password,
                    "user_id": str(user.id),
                    "center": center.name,
                },
                status=status.HTTP_201_CREATED,
            )
    except Exception as e:
        return Response(
            {"detail": f"Error creating admin: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


def _get_center_by_name_or_id(center_name=None, center_id=None):
    """
    Helper function to find a center by name or ID.
    Returns (center, error_response) tuple.
    If center is found, returns (center, None).
    If error, returns (None, error_response).
    """
    center = None
    
    if center_id:
        try:
            center = Center.objects.get(id=center_id)
            return center, None
        except Center.DoesNotExist:
            return None, Response(
                {"detail": f"Center with id '{center_id}' not found."},
                status=status.HTTP_404_NOT_FOUND,
            )
    elif center_name:
        centers = Center.objects.filter(name=center_name)
        if centers.count() == 0:
            return None, Response(
                {"detail": f"Center '{center_name}' not found. Please create the center first."},
                status=status.HTTP_404_NOT_FOUND,
            )
        elif centers.count() > 1:
            # Return list of matching centers so user can choose by ID
            centers_list = [
                {"id": str(c.id), "name": c.name, "city": c.city, "institute": c.institute.name}
                for c in centers
            ]
            return None, Response(
                {
                    "detail": f"Multiple centers found with name '{center_name}'. Please use center_id to specify which one.",
                    "matching_centers": centers_list
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        else:
            center = centers.first()
            return center, None
    
    return None, Response(
        {"detail": "Either center_name or center_id is required."},
        status=status.HTTP_400_BAD_REQUEST,
    )


def _check_admin_or_super(request):
    """Helper to check if user is admin or super admin."""
    if not request.user.is_authenticated:
        return False, None, Response(
            {"detail": "Authentication required."},
            status=status.HTTP_401_UNAUTHORIZED,
        )
    user = request.user
    user_role = user.role.lower() if user.role else ''
    is_super = user_role in ['super_admin', 'SUPER_ADMIN']
    is_admin = user_role in ['admin', 'institute_admin', 'exam_admin']
    
    if not (is_super or is_admin):
        return False, None, Response(
            {"detail": f"Only Admin or Super Admin can perform this action. Your role: {user.role}"},
            status=status.HTTP_403_FORBIDDEN,
        )
    
    return True, (is_super, is_admin), None


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_teacher(request):
    """
    Super Admin or Admin creates a new Teacher user.
    
    - Super Admin: Can create teacher in any center (must provide center_name)
    - Admin: Can create teacher in their own center (center_name optional)
    
    Auto-generates:
    - username (code): TCH-<center_code>-<3_digits>
    - password: Teacher@<center_code><year>
    - teacher_code: Same as username
    
    Payload:
    {
        "center_name": "Allen - Jaipur Center",  # Required for Super Admin, optional for Admin
        "name": "Teacher Name",
        "email": "teacher@example.com",
        "phone_number": "9876543210",
        "employee_id": "EMP-001",
        "subjects": "Physics, Chemistry"
    }
    
    Returns:
    {
        "username": "TCH-XXXX-123",
        "password": "Teacher@XXXX2025",
        "teacher_code": "TCH-XXXX-123",
        "user_id": "uuid"
    }
    """
    can_proceed, role_info, error_response = _check_admin_or_super(request)
    if not can_proceed:
        return error_response
    
    is_super, is_admin = role_info
    user = request.user
    
    center_name = request.data.get("center_name")
    center_id = request.data.get("center_id")
    name = request.data.get("name")
    email = request.data.get("email", "")
    phone_number = request.data.get("phone_number", "")
    employee_id = request.data.get("employee_id", "")
    subjects = request.data.get("subjects", "")
    
    if not name:
        return Response(
            {"detail": "name is required."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    # Determine center
    if is_super:
        # Super Admin must provide center_name or center_id
        if not center_name and not center_id:
            return Response(
                {"detail": "Either center_name or center_id is required for Super Admin."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center, error_response = _get_center_by_name_or_id(center_name=center_name, center_id=center_id)
        if error_response:
            return error_response
    else:
        # Admin uses their own center
        if not user.center:
            return Response(
                {"detail": "Admin user is not linked to any center."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center = user.center
    
    # Generate code and password
    center_code = center.name[:4].replace(" ", "").replace("-", "")
    username = generate_user_code('TEACHER', center_code)
    password = generate_password('TEACHER', center_code)
    teacher_code = username  # Use same code
    
    # Split name
    name_parts = name.strip().split()
    first_name = name_parts[0] if name_parts else name
    last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
    
    try:
        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                email=email or f"{username}@temp.com",
                password=password,
                first_name=first_name,
                last_name=last_name,
                phone=phone_number or "",
                phone_number=phone_number or "",
                role='teacher',  # Use exam role format (compatible with both)
                center=center,
                institute=center.institute,
                teacher_code=teacher_code,
                teacher_employee_id=employee_id,
                teacher_subjects=subjects,
            )
            
            return Response(
                {
                    "message": "Teacher created successfully.",
                    "username": username,
                    "password": password,
                    "teacher_code": teacher_code,
                    "user_id": str(user.id),
                    "center": center.name,
                },
                status=status.HTTP_201_CREATED,
            )
    except Exception as e:
        return Response(
            {"detail": f"Error creating teacher: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_student(request):
    """
    Super Admin or Admin creates a new Student user.
    
    - Super Admin: Can create student in any batch
    - Admin: Can create student in batches from their center
    
    Auto-generates:
    - username (code): STU-<batch_code>-<4_digits>
    - password: Student@<batch_code><year> or Student@<dob_year>
    
    Payload:
    {
        "batch_code": "HDTN-1A-ZA1",
        "name": "Student Name",
        "email": "student@example.com",
        "phone_number": "9876543210",
        "date_of_birth": "2010-05-15"
    }
    
    Returns:
    {
        "username": "STU-XXXX-1234",
        "password": "Student@XXXX2025",
        "user_id": "uuid"
    }
    """
    can_proceed, role_info, error_response = _check_admin_or_super(request)
    if not can_proceed:
        return error_response
    
    is_super, is_admin = role_info
    user = request.user
    
    batch_code = request.data.get("batch_code")
    name = request.data.get("name")
    email = request.data.get("email", "")
    phone_number = request.data.get("phone_number", "")
    date_of_birth = request.data.get("date_of_birth", "")
    
    if not batch_code or not name:
        return Response(
            {"detail": "batch_code and name are required."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    try:
        batch = Batch.objects.select_related('program', 'program__center').get(code=batch_code)
        center = batch.program.center
    except Batch.DoesNotExist:
        return Response(
            {"detail": f"Batch with code '{batch_code}' not found. Please create the batch first."},
            status=status.HTTP_404_NOT_FOUND,
        )
    except Batch.MultipleObjectsReturned:
        return Response(
            {"detail": f"Multiple batches found with code '{batch_code}'. Please use batch_id instead."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    # Check permissions: Admin can only create students in their center's batches
    if is_admin:
        if not user.center:
            return Response(
                {"detail": "Admin user is not linked to any center."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if center != user.center:
            return Response(
                {"detail": f"You can only create students in batches from your center '{user.center.name}'."},
                status=status.HTTP_403_FORBIDDEN,
            )
    
    # Generate code and password
    username = generate_user_code('STUDENT', None, batch_code)
    password = generate_password('STUDENT', None, batch_code, date_of_birth)
    
    # Split name
    name_parts = name.strip().split()
    first_name = name_parts[0] if name_parts else name
    last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
    
    try:
        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                email=email or f"{username}@temp.com",
                password=password,
                first_name=first_name,
                last_name=last_name,
                phone=phone_number or "",
                phone_number=phone_number or "",
                role='student',  # Use exam role format (compatible with both)
                center=center,
                institute=center.institute,
            )
            
            # Auto-enroll student in the batch
            Enrollment.objects.get_or_create(
                student=user,
                batch=batch,
                defaults={'status': Enrollment.STATUS_ACTIVE}
            )
            
            return Response(
                {
                    "message": "Student created successfully.",
                    "username": username,
                    "password": password,
                    "user_id": str(user.id),
                    "batch": batch.code,
                    "center": center.name,
                },
                status=status.HTTP_201_CREATED,
            )
    except Exception as e:
        return Response(
            {"detail": f"Error creating student: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_staff(request):
    """
    Super Admin or Admin creates a new Staff user.
    
    - Super Admin: Can create staff in any center (must provide center_name)
    - Admin: Can create staff in their own center (center_name optional)
    
    Auto-generates:
    - username (code): STF-<center_code>-<3_digits>
    - password: Staff@<center_code><year>
    
    Payload:
    {
        "center_name": "Allen - Jaipur Center",  # Required for Super Admin, optional for Admin
        "name": "Staff Name",
        "email": "staff@example.com",
        "phone_number": "9876543210"
    }
    
    Returns:
    {
        "username": "STF-XXXX-123",
        "password": "Staff@XXXX2025",
        "user_id": "uuid"
    }
    """
    can_proceed, role_info, error_response = _check_admin_or_super(request)
    if not can_proceed:
        return error_response
    
    is_super, is_admin = role_info
    user = request.user
    
    center_name = request.data.get("center_name")
    center_id = request.data.get("center_id")
    name = request.data.get("name")
    email = request.data.get("email", "")
    phone_number = request.data.get("phone_number", "")
    
    if not name:
        return Response(
            {"detail": "name is required."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    # Determine center
    if is_super:
        # Super Admin must provide center_name or center_id
        if not center_name and not center_id:
            return Response(
                {"detail": "Either center_name or center_id is required for Super Admin."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center, error_response = _get_center_by_name_or_id(center_name=center_name, center_id=center_id)
        if error_response:
            return error_response
    else:
        # Admin uses their own center
        if not user.center:
            return Response(
                {"detail": "Admin user is not linked to any center."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center = user.center
    
    # Generate code and password
    center_code = center.name[:4].replace(" ", "").replace("-", "")
    username = generate_user_code('STAFF', center_code)
    password = generate_password('STAFF', center_code)
    
    # Split name
    name_parts = name.strip().split()
    first_name = name_parts[0] if name_parts else name
    last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
    
    try:
        with transaction.atomic():
            user_obj = User.objects.create_user(
                username=username,
                email=email or f"{username}@temp.com",
                password=password,
                first_name=first_name,
                last_name=last_name,
                phone=phone_number or "",
                phone_number=phone_number or "",
                role='staff',  # Use lowercase staff role
                center=center,
                institute=center.institute,
            )
            
            return Response(
                {
                    "message": "Staff created successfully.",
                    "username": username,
                    "password": password,
                    "user_id": str(user_obj.id),
                    "center": center.name,
                },
                status=status.HTTP_201_CREATED,
            )
    except Exception as e:
        return Response(
            {"detail": f"Error creating staff: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def bulk_create_teachers(request):
    """
    Bulk create teachers from Excel/CSV file or JSON array.
    
    - Super Admin: Can create teachers in any center (must provide center_id)
    - Admin: Can create teachers in their own center
    
    Accepts either:
    1. File upload (Excel .xlsx or CSV) with columns: name, email, phone_number, employee_id, subjects
    2. JSON array in request body with key "teachers"
    
    File Upload:
    - Content-Type: multipart/form-data
    - file: Excel or CSV file
    - center_id: (optional for Admin, required for Super Admin)
    
    JSON Payload:
    {
        "center_id": "uuid",  # Optional for Admin
        "teachers": [
            {
                "name": "Teacher Name",
                "email": "teacher@example.com",
                "phone_number": "9876543210",
                "employee_id": "EMP-001",
                "subjects": "Physics, Chemistry"
            },
            ...
        ]
    }
    
    Returns:
    {
        "message": "Bulk teacher creation completed.",
        "total": 10,
        "success": 8,
        "failed": 2,
        "created_teachers": [...],
        "errors": [...]
    }
    """
    import openpyxl
    import csv
    import io
    
    can_proceed, role_info, error_response = _check_admin_or_super(request)
    if not can_proceed:
        return error_response
    
    is_super, is_admin = role_info
    user = request.user
    
    # Get center
    center_id = request.data.get("center_id") or request.POST.get("center_id")
    
    if is_super:
        if not center_id:
            return Response(
                {"detail": "center_id is required for Super Admin."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center, error_response = _get_center_by_name_or_id(center_id=center_id)
        if error_response:
            return error_response
    else:
        if not user.center:
            return Response(
                {"detail": "Admin user is not linked to any center."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center = user.center
    
    teachers_data = []
    
    # Check if file upload
    if 'file' in request.FILES:
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        
        try:
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                # Parse Excel file
                wb = openpyxl.load_workbook(uploaded_file, read_only=True)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).lower().strip() if cell.value else '')
                
                # Map headers to expected fields
                header_map = {}
                for i, h in enumerate(headers):
                    if 'name' in h and 'employee' not in h:
                        header_map['name'] = i
                    elif 'email' in h:
                        header_map['email'] = i
                    elif 'phone' in h:
                        header_map['phone_number'] = i
                    elif 'employee' in h or 'emp' in h:
                        header_map['employee_id'] = i
                    elif 'subject' in h:
                        header_map['subjects'] = i
                
                if 'name' not in header_map:
                    return Response(
                        {"detail": "Excel file must have a 'name' column."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                
                # Parse rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    teacher = {
                        'name': str(row[header_map.get('name', 0)] or '').strip(),
                        'email': str(row[header_map.get('email', 1)] or '').strip() if header_map.get('email') is not None and len(row) > header_map.get('email', 1) else '',
                        'phone_number': str(row[header_map.get('phone_number', 2)] or '').strip() if header_map.get('phone_number') is not None and len(row) > header_map.get('phone_number', 2) else '',
                        'employee_id': str(row[header_map.get('employee_id', 3)] or '').strip() if header_map.get('employee_id') is not None and len(row) > header_map.get('employee_id', 3) else '',
                        'subjects': str(row[header_map.get('subjects', 4)] or '').strip() if header_map.get('subjects') is not None and len(row) > header_map.get('subjects', 4) else '',
                        'row': row_idx,
                    }
                    
                    if teacher['name']:
                        teachers_data.append(teacher)
                
                wb.close()
                
            elif file_name.endswith('.csv'):
                # Parse CSV file
                content = uploaded_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(content))
                
                for row_idx, row in enumerate(reader, start=2):
                    # Normalize keys
                    normalized = {}
                    for k, v in row.items():
                        key = k.lower().strip()
                        if 'name' in key and 'employee' not in key:
                            normalized['name'] = v
                        elif 'email' in key:
                            normalized['email'] = v
                        elif 'phone' in key:
                            normalized['phone_number'] = v
                        elif 'employee' in key or 'emp' in key:
                            normalized['employee_id'] = v
                        elif 'subject' in key:
                            normalized['subjects'] = v
                    
                    teacher = {
                        'name': str(normalized.get('name', '') or '').strip(),
                        'email': str(normalized.get('email', '') or '').strip(),
                        'phone_number': str(normalized.get('phone_number', '') or '').strip(),
                        'employee_id': str(normalized.get('employee_id', '') or '').strip(),
                        'subjects': str(normalized.get('subjects', '') or '').strip(),
                        'row': row_idx,
                    }
                    
                    if teacher['name']:
                        teachers_data.append(teacher)
            else:
                return Response(
                    {"detail": "Unsupported file format. Please upload .xlsx or .csv file."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response(
                {"detail": f"Error parsing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    else:
        # JSON payload
        teachers_data = request.data.get("teachers", [])
        if not teachers_data:
            return Response(
                {"detail": "No teachers data provided. Upload a file or provide 'teachers' array."},
                status=status.HTTP_400_BAD_REQUEST,
            )
    
    if not teachers_data:
        return Response(
            {"detail": "No valid teacher records found in the uploaded data."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    # Process teachers
    created_teachers = []
    errors = []
    center_code = center.name[:4].replace(" ", "").replace("-", "")
    
    for idx, teacher_data in enumerate(teachers_data):
        name = teacher_data.get('name', '').strip()
        email = teacher_data.get('email', '').strip()
        phone_number = teacher_data.get('phone_number', '').strip()
        employee_id = teacher_data.get('employee_id', '').strip()
        subjects = teacher_data.get('subjects', '').strip()
        row_num = teacher_data.get('row', idx + 1)
        
        if not name:
            errors.append({
                'row': row_num,
                'error': 'Name is required',
                'data': teacher_data
            })
            continue
        
        # Check for duplicate email
        if email and User.objects.filter(email=email).exists():
            errors.append({
                'row': row_num,
                'error': f'Email {email} already exists',
                'data': teacher_data
            })
            continue
        
        try:
            # Generate code and password
            username = generate_user_code('TEACHER', center_code)
            password = generate_password('TEACHER', center_code)
            teacher_code = username
            
            # Split name
            name_parts = name.strip().split()
            first_name = name_parts[0] if name_parts else name
            last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
            
            with transaction.atomic():
                new_user = User.objects.create_user(
                    username=username,
                    email=email or f"{username}@temp.com",
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone_number or "",
                    phone_number=phone_number or "",
                    role='teacher',
                    center=center,
                    institute=center.institute,
                    teacher_code=teacher_code,
                    teacher_employee_id=employee_id,
                    teacher_subjects=subjects,
                )
                
                created_teachers.append({
                    'row': row_num,
                    'name': name,
                    'username': username,
                    'password': password,
                    'teacher_code': teacher_code,
                    'email': new_user.email,
                    'user_id': str(new_user.id),
                })
        except Exception as e:
            errors.append({
                'row': row_num,
                'error': str(e),
                'data': teacher_data
            })
    
    return Response({
        'message': 'Bulk teacher creation completed.',
        'total': len(teachers_data),
        'success': len(created_teachers),
        'failed': len(errors),
        'created_teachers': created_teachers,
        'errors': errors,
        'center': center.name,
    }, status=status.HTTP_201_CREATED if created_teachers else status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def bulk_create_students(request):
    """
    Bulk create students from Excel/CSV file or JSON array.
    
    Accepts either:
    1. File upload (Excel .xlsx or CSV) with columns: name, email, phone_number, batch_code, date_of_birth
    2. JSON array in request body with key "students"
    
    File Upload:
    - Content-Type: multipart/form-data
    - file: Excel or CSV file
    - center_id: (optional for Admin, required for Super Admin)
    
    Returns:
    {
        "message": "Bulk student creation completed.",
        "total": 10,
        "success": 8,
        "failed": 2,
        "created_students": [...],
        "errors": [...]
    }
    """
    import openpyxl
    import csv
    import io
    
    can_proceed, role_info, error_response = _check_admin_or_super(request)
    if not can_proceed:
        return error_response
    
    is_super, is_admin = role_info
    user = request.user
    
    # Get center
    center_id = request.data.get("center_id") or request.POST.get("center_id")
    
    if is_super:
        if not center_id:
            return Response(
                {"detail": "center_id is required for Super Admin."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center, error_response = _get_center_by_name_or_id(center_id=center_id)
        if error_response:
            return error_response
    else:
        if not user.center:
            return Response(
                {"detail": "Admin user is not linked to any center."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center = user.center
    
    students_data = []
    
    # Check if file upload
    if 'file' in request.FILES:
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        
        try:
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                # Parse Excel file
                wb = openpyxl.load_workbook(uploaded_file, read_only=True)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).lower().strip() if cell.value else '')
                
                # Map headers to expected fields
                header_map = {}
                for i, h in enumerate(headers):
                    if 'name' in h and 'batch' not in h:
                        header_map['name'] = i
                    elif 'email' in h:
                        header_map['email'] = i
                    elif 'phone' in h:
                        header_map['phone_number'] = i
                    elif 'batch' in h:
                        header_map['batch_code'] = i
                    elif 'birth' in h or 'dob' in h:
                        header_map['date_of_birth'] = i
                
                if 'name' not in header_map:
                    return Response(
                        {"detail": "Excel file must have a 'name' column."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                
                # Parse rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    student = {
                        'name': str(row[header_map.get('name', 0)] or '').strip(),
                        'email': str(row[header_map.get('email', 1)] or '').strip() if header_map.get('email') is not None and len(row) > header_map.get('email', 1) else '',
                        'phone_number': str(row[header_map.get('phone_number', 2)] or '').strip() if header_map.get('phone_number') is not None and len(row) > header_map.get('phone_number', 2) else '',
                        'batch_code': str(row[header_map.get('batch_code', 3)] or '').strip() if header_map.get('batch_code') is not None and len(row) > header_map.get('batch_code', 3) else '',
                        'date_of_birth': str(row[header_map.get('date_of_birth', 4)] or '').strip() if header_map.get('date_of_birth') is not None and len(row) > header_map.get('date_of_birth', 4) else '',
                        'row': row_idx,
                    }
                    
                    if student['name']:
                        students_data.append(student)
                
                wb.close()
                
            elif file_name.endswith('.csv'):
                # Parse CSV file
                content = uploaded_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(content))
                
                for row_idx, row in enumerate(reader, start=2):
                    # Normalize keys
                    normalized = {}
                    for k, v in row.items():
                        key = k.lower().strip()
                        if 'name' in key and 'batch' not in key:
                            normalized['name'] = v
                        elif 'email' in key:
                            normalized['email'] = v
                        elif 'phone' in key:
                            normalized['phone_number'] = v
                        elif 'batch' in key:
                            normalized['batch_code'] = v
                        elif 'birth' in key or 'dob' in key:
                            normalized['date_of_birth'] = v
                    
                    student = {
                        'name': str(normalized.get('name', '') or '').strip(),
                        'email': str(normalized.get('email', '') or '').strip(),
                        'phone_number': str(normalized.get('phone_number', '') or '').strip(),
                        'batch_code': str(normalized.get('batch_code', '') or '').strip(),
                        'date_of_birth': str(normalized.get('date_of_birth', '') or '').strip(),
                        'row': row_idx,
                    }
                    
                    if student['name']:
                        students_data.append(student)
            else:
                return Response(
                    {"detail": "Unsupported file format. Please upload .xlsx or .csv file."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response(
                {"detail": f"Error parsing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    else:
        # JSON payload
        students_data = request.data.get("students", [])
        if not students_data:
            return Response(
                {"detail": "No students data provided. Upload a file or provide 'students' array."},
                status=status.HTTP_400_BAD_REQUEST,
            )
    
    if not students_data:
        return Response(
            {"detail": "No valid student records found in the uploaded data."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    # Process students
    created_students = []
    errors = []
    center_code = center.name[:4].replace(" ", "").replace("-", "")
    
    for idx, student_data in enumerate(students_data):
        name = student_data.get('name', '').strip()
        email = student_data.get('email', '').strip()
        phone_number = student_data.get('phone_number', '').strip()
        batch_code = student_data.get('batch_code', '').strip()
        date_of_birth = student_data.get('date_of_birth', '').strip()
        row_num = student_data.get('row', idx + 1)
        
        if not name:
            errors.append({
                'row': row_num,
                'error': 'Name is required',
                'data': student_data
            })
            continue
        
        # Check for duplicate email
        if email and User.objects.filter(email=email).exists():
            errors.append({
                'row': row_num,
                'error': f'Email {email} already exists',
                'data': student_data
            })
            continue
        
        try:
            # Generate code and password
            username = generate_user_code('STUDENT', center_code)
            password = generate_password('STUDENT', center_code)
            
            # Split name
            name_parts = name.strip().split()
            first_name = name_parts[0] if name_parts else name
            last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
            
            with transaction.atomic():
                new_user = User.objects.create_user(
                    username=username,
                    email=email or f"{username}@temp.com",
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone_number or "",
                    phone_number=phone_number or "",
                    role='student',
                    center=center,
                    institute=center.institute,
                )
                
                # If batch_code is provided, try to enroll student in that batch
                if batch_code:
                    try:
                        batch = Batch.objects.get(code=batch_code)
                        Enrollment.objects.get_or_create(
                            student=new_user,
                            batch=batch,
                            defaults={'status': Enrollment.STATUS_ACTIVE}
                        )
                    except Batch.DoesNotExist:
                        pass  # Batch not found, skip enrollment
                
                created_students.append({
                    'row': row_num,
                    'name': name,
                    'username': username,
                    'password': password,
                    'batch_code': batch_code,
                    'email': new_user.email,
                    'user_id': str(new_user.id),
                })
        except Exception as e:
            errors.append({
                'row': row_num,
                'error': str(e),
                'data': student_data
            })
    
    return Response({
        'message': 'Bulk student creation completed.',
        'total': len(students_data),
        'success': len(created_students),
        'failed': len(errors),
        'created_students': created_students,
        'errors': errors,
        'center': center.name,
    }, status=status.HTTP_201_CREATED if created_students else status.HTTP_400_BAD_REQUEST)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def bulk_create_staff(request):
    """
    Bulk create staff from Excel/CSV file or JSON array.
    
    Accepts either:
    1. File upload (Excel .xlsx or CSV) with columns: name, email, phone_number
    2. JSON array in request body with key "staff"
    
    File Upload:
    - Content-Type: multipart/form-data
    - file: Excel or CSV file
    - center_id: (optional for Admin, required for Super Admin)
    
    Returns:
    {
        "message": "Bulk staff creation completed.",
        "total": 10,
        "success": 8,
        "failed": 2,
        "created_staff": [...],
        "errors": [...]
    }
    """
    import openpyxl
    import csv
    import io
    
    can_proceed, role_info, error_response = _check_admin_or_super(request)
    if not can_proceed:
        return error_response
    
    is_super, is_admin = role_info
    user = request.user
    
    # Get center
    center_id = request.data.get("center_id") or request.POST.get("center_id")
    
    if is_super:
        if not center_id:
            return Response(
                {"detail": "center_id is required for Super Admin."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center, error_response = _get_center_by_name_or_id(center_id=center_id)
        if error_response:
            return error_response
    else:
        if not user.center:
            return Response(
                {"detail": "Admin user is not linked to any center."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        center = user.center
    
    staff_data = []
    
    # Check if file upload
    if 'file' in request.FILES:
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        
        try:
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                # Parse Excel file
                wb = openpyxl.load_workbook(uploaded_file, read_only=True)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value).lower().strip() if cell.value else '')
                
                # Map headers to expected fields
                header_map = {}
                for i, h in enumerate(headers):
                    if 'name' in h:
                        header_map['name'] = i
                    elif 'email' in h:
                        header_map['email'] = i
                    elif 'phone' in h:
                        header_map['phone_number'] = i
                
                if 'name' not in header_map:
                    return Response(
                        {"detail": "Excel file must have a 'name' column."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                
                # Parse rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not any(row):
                        continue
                    
                    staff = {
                        'name': str(row[header_map.get('name', 0)] or '').strip(),
                        'email': str(row[header_map.get('email', 1)] or '').strip() if header_map.get('email') is not None and len(row) > header_map.get('email', 1) else '',
                        'phone_number': str(row[header_map.get('phone_number', 2)] or '').strip() if header_map.get('phone_number') is not None and len(row) > header_map.get('phone_number', 2) else '',
                        'row': row_idx,
                    }
                    
                    if staff['name']:
                        staff_data.append(staff)
                
                wb.close()
                
            elif file_name.endswith('.csv'):
                # Parse CSV file
                content = uploaded_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(content))
                
                for row_idx, row in enumerate(reader, start=2):
                    # Normalize keys
                    normalized = {}
                    for k, v in row.items():
                        key = k.lower().strip()
                        if 'name' in key:
                            normalized['name'] = v
                        elif 'email' in key:
                            normalized['email'] = v
                        elif 'phone' in key:
                            normalized['phone_number'] = v
                    
                    staff = {
                        'name': str(normalized.get('name', '') or '').strip(),
                        'email': str(normalized.get('email', '') or '').strip(),
                        'phone_number': str(normalized.get('phone_number', '') or '').strip(),
                        'row': row_idx,
                    }
                    
                    if staff['name']:
                        staff_data.append(staff)
            else:
                return Response(
                    {"detail": "Unsupported file format. Please upload .xlsx or .csv file."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response(
                {"detail": f"Error parsing file: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    else:
        # JSON payload
        staff_data = request.data.get("staff", [])
        if not staff_data:
            return Response(
                {"detail": "No staff data provided. Upload a file or provide 'staff' array."},
                status=status.HTTP_400_BAD_REQUEST,
            )
    
    if not staff_data:
        return Response(
            {"detail": "No valid staff records found in the uploaded data."},
            status=status.HTTP_400_BAD_REQUEST,
        )
    
    # Process staff
    created_staff = []
    errors = []
    center_code = center.name[:4].replace(" ", "").replace("-", "")
    
    for idx, staff_member in enumerate(staff_data):
        name = staff_member.get('name', '').strip()
        email = staff_member.get('email', '').strip()
        phone_number = staff_member.get('phone_number', '').strip()
        row_num = staff_member.get('row', idx + 1)
        
        if not name:
            errors.append({
                'row': row_num,
                'error': 'Name is required',
                'data': staff_member
            })
            continue
        
        # Check for duplicate email
        if email and User.objects.filter(email=email).exists():
            errors.append({
                'row': row_num,
                'error': f'Email {email} already exists',
                'data': staff_member
            })
            continue
        
        try:
            # Generate code and password
            username = generate_user_code('STAFF', center_code)
            password = generate_password('STAFF', center_code)
            
            # Split name
            name_parts = name.strip().split()
            first_name = name_parts[0] if name_parts else name
            last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
            
            with transaction.atomic():
                new_user = User.objects.create_user(
                    username=username,
                    email=email or f"{username}@temp.com",
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    phone=phone_number or "",
                    phone_number=phone_number or "",
                    role=User.ROLE_STAFF,  # Use proper role constant
                    center=center,
                    institute=center.institute,
                )
                
                created_staff.append({
                    'row': row_num,
                    'name': name,
                    'username': username,
                    'password': password,
                    'email': new_user.email,
                    'user_id': str(new_user.id),
                })
        except Exception as e:
            errors.append({
                'row': row_num,
                'error': str(e),
                'data': staff_member
            })
    
    return Response({
        'message': 'Bulk staff creation completed.',
        'total': len(staff_data),
        'success': len(created_staff),
        'failed': len(errors),
        'created_staff': created_staff,
        'errors': errors,
        'center': center.name,
    }, status=status.HTTP_201_CREATED if created_staff else status.HTTP_400_BAD_REQUEST)

